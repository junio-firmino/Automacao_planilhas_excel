from openpyxl import load_workbook
import datetime as dt
from dateutil.relativedelta import relativedelta
import smtplib
from email.message import EmailMessage
from locale import setlocale, LC_ALL


class Risco:
    def __init__(self):
        self.cliente = 'utf-8'
        self.taxas = 0
        self.wb = load_workbook(filename='template_risco_sacado.xlsx')
        self.wb_cpgt = load_workbook(filename='template_cpgt_risco_sacado.xlsx')

    def interface(self):
        self.abrir_arq()
        active = True
        while active:
            self.cliente = input('Qual cliente você irá cadastrar? ').title()
            self.taxas = input('Qual a taxa? ')
            self.cpgt = input('Qual o prazo da condição de pagamento? ')
            self.banco = self.banco_1()
            self.lista_distr()
            self.abrir_plan_risco_sacado()
            self.abrir_plan_cpgt()
            self.salvar_arq()
            alerta = input('Prosseguir o cadastro ? \n(Pressione "enter" para continuar com os cadastros.\n'
                           'Caso deseje finalizar pressione "f" em seguida "enter".)-->')

            if alerta == 'f':
                active = False
        self.enviar_email()

    def abrir_arq(self):
        return self.wb and self.wb_cpgt

    def salvar_arq(self):
        self.wb.save('risco_sacado(' + self.data_save_arquivo() + ').xlsx')
        self.wb_cpgt.save('Cadastro_CPGT_RS(' + self.data_save_arquivo() + ').xlsx')

    def cpgt_terrestre(self):
        return 'ZD'+self.cpgt

    def cpgt_cabotagem(self):
        return 'ZC'+self.cpgt

    @staticmethod
    def banco_1():
        flag = True
        while flag:
            bancos = {'s': 'Santander', 'b': 'Bradesco', 'c': 'Citibank'}
            banco_marca = input('Escolha o banco?\n("s" para Santander, "b" para Bradesco e "c" para Citibank)'
                                ' + "enter" -->')
            if banco_marca in bancos:
                return bancos[banco_marca]
            else:
                print('Essa escolha não é possível, tente novamente!.')

    def lista_distr(self):
        distri = [self.cliente, self.taxas, self.cpgt_terrestre(), self.cpgt_cabotagem(), self.banco]
        return distri

    @staticmethod
    def centro_terrestre():
        terrestre = [1700, 1400, 1200, 1210, 1100, 1360, 1950, 1101, 1110, 1111, 1120, 1130]
        return terrestre

    @staticmethod
    def centro_cabotagem():
        cabotagem = [1410]
        return cabotagem

    def abrir_plan_risco_sacado(self):
        aba_act = self.wb.active
        self.lista_distr()
        for linha_plan in range(2, aba_act.max_row + 1):  # Tratamento na planilha das linhas
            empresa = aba_act.cell(row=linha_plan, column=2).value
            centro_1 = aba_act.cell(row=linha_plan, column=6).value
            if empresa in self.lista_distr() and centro_1 in self.centro_terrestre():
                aba_act.cell(row=linha_plan, column=8).value = self.lista_distr()[2]              # Preenche CPGT
                aba_act.cell(row=linha_plan, column=9).value = self.lista_distr()[1]+" a.m"       # Preenche taxa
                aba_act.cell(row=linha_plan, column=10).value = self.data_inicio()                # Preenche data inicio
                aba_act.cell(row=linha_plan, column=11).value = self.data_last_day_risco_sacado()  # Preenche data final
                aba_act.cell(row=linha_plan, column=12).value = self.lista_distr()[-1]            # Preenche Banco
                aba_act.cell(row=linha_plan, column=13).value = self.data_cadastro()
            elif empresa in self.lista_distr() and centro_1 in self.centro_cabotagem():
                aba_act.cell(row=linha_plan, column=8).value = self.lista_distr()[3]
                aba_act.cell(row=linha_plan, column=9).value = self.lista_distr()[1]+" a.m"
                aba_act.cell(row=linha_plan, column=10).value = self.data_inicio()
                aba_act.cell(row=linha_plan, column=11).value = self.data_last_day_risco_sacado()
                aba_act.cell(row=linha_plan, column=12).value = self.lista_distr()[-1]
                aba_act.cell(row=linha_plan, column=13).value = self.data_cadastro()

    def abrir_plan_cpgt(self):
        aba_act_cpgt = self.wb_cpgt.active
        self.lista_distr()
        for linha_cpgt in range(2, aba_act_cpgt.max_row + 1):
            distribuidora = aba_act_cpgt.cell(row=linha_cpgt, column=19).value
            centro_cpgt = aba_act_cpgt.cell(row=linha_cpgt, column=8).value
            if distribuidora in self.lista_distr() and centro_cpgt in self.centro_terrestre():
                aba_act_cpgt.cell(row=linha_cpgt, column=3).value = self.lista_distr()[2]
                aba_act_cpgt.cell(row=linha_cpgt, column=16).value = self.data_inicio()
                aba_act_cpgt.cell(row=linha_cpgt, column=17).value = self.data_last_day_cpgt()
                aba_act_cpgt.cell(row=linha_cpgt, column=7).value = self.carencia_cpgt_terrestre()
            elif distribuidora in self.lista_distr() and centro_cpgt in self.centro_cabotagem():
                aba_act_cpgt.cell(row=linha_cpgt, column=3).value = self.lista_distr()[3]
                aba_act_cpgt.cell(row=linha_cpgt, column=16).value = self.data_inicio()
                aba_act_cpgt.cell(row=linha_cpgt, column=17).value = self.data_last_day_cpgt()
                aba_act_cpgt.cell(row=linha_cpgt, column=7).value = self.carencia_cpgt_cabotagem()

    def carencia_cpgt_terrestre(self):
        condicoes_cpgt = self.lista_distr()[2]
        list_separador = condicoes_cpgt.split('D')
        valor_separado = list_separador[1]
        resultado = int(valor_separado) - 1
        return resultado

    def carencia_cpgt_cabotagem(self):
        condicoes_cpgt_cabotagem = self.lista_distr()[3]
        list_separador_cabotagem = condicoes_cpgt_cabotagem.split('C')
        valor_separado_cabotagem = list_separador_cabotagem[1]
        resultado_cabotagem = int(valor_separado_cabotagem) - 4
        return resultado_cabotagem

    @staticmethod
    def data_save_arquivo():
        data_save = dt.datetime.now()
        return data_save.strftime('%d_%m_%y')

    @staticmethod
    def data_cadastro():
        data_cad = dt.datetime.now()
        return data_cad.strftime('%d.%m.%Y')

    def data_inicio(self):
        data_ini = dt.datetime.now() + relativedelta(months=1)
        data_1 = dt.datetime.now().strftime('01.%m.%Y')
        data_1_transforma_date = dt.datetime.strptime(data_1, '%d.%m.%Y')
        data_2 = dt.datetime.now().strftime('21.%m.%Y')
        data_2_transforma_date = dt.datetime.strptime(data_2, '%d.%m.%Y')
        data_3 = dt.datetime.now().strftime('%d.%m.%Y')
        data_3_transforma_date = dt.datetime.strptime(data_3, '%d.%m.%Y')
        data_1_day = data_1_transforma_date.day
        data_2_day = data_2_transforma_date.day
        data_3_day = data_3_transforma_date.day
        if data_3_day in range(0, (data_2_day - data_1_day)):
            return self.data_cadastro()
        else:
            return data_ini.strftime('01.%m.%Y')

    @staticmethod
    def data_last_day_risco_sacado():
        data_last = dt.datetime.now() + relativedelta(day=31, months=1)
        data_last_1 = dt.datetime.now() + relativedelta(day=31)
        data_last_1_1 = dt.datetime.now().strftime('01.%m.%Y')
        data_last_1_1_transforma_date = dt.datetime.strptime(data_last_1_1, '%d.%m.%Y')
        data_last_2 = dt.datetime.now().strftime('21.%m.%Y')
        data_last_2_transforma_date = dt.datetime.strptime(data_last_2, '%d.%m.%Y')
        data_last_3 = dt.datetime.now().strftime('%d.%m.%Y')
        data_last_3_transforma_date = dt.datetime.strptime(data_last_3, '%d.%m.%Y')
        data_last_1_day = data_last_1_1_transforma_date.day
        data_last_2_day = data_last_2_transforma_date.day
        data_last_3_day = data_last_3_transforma_date.day
        if data_last_3_day in range(0, (data_last_2_day - data_last_1_day)):
            return data_last_1.strftime('%d.%m.%Y')
        else:
            return data_last.strftime('%d.%m.%Y')

    @staticmethod
    def data_last_day_cpgt():
        data_last_cpgt = dt.datetime.now() + relativedelta(day=1, months=3)
        data_last_1_cpgt = dt.datetime.now() + relativedelta(day=1, months=2)
        data_last_1_1_cpgt = dt.datetime.now().strftime('01.%m.%Y')
        data_last_1_1_cpgt_transforma_date = dt.datetime.strptime(data_last_1_1_cpgt, '%d.%m.%Y')
        data_last_2_cpgt = dt.datetime.now().strftime('21.%m.%Y')
        data_last_2_cpgt_transforma_date = dt.datetime.strptime(data_last_2_cpgt, '%d.%m.%Y')
        data_last_3_cpgt = dt.datetime.now().strftime('%d.%m.%Y')
        data_last_3_cpgt_transforma_date = dt.datetime.strptime(data_last_3_cpgt, '%d.%m.%Y')
        data_last_1_cpgt_day = data_last_1_1_cpgt_transforma_date.day
        data_last_2_cpgt_day = data_last_2_cpgt_transforma_date.day
        data_last_3_cpgt_day = data_last_3_cpgt_transforma_date.day
        if data_last_3_cpgt_day in range(0, (data_last_2_cpgt_day - data_last_1_cpgt_day)):
            return data_last_1_cpgt.strftime('%d.%m.%Y')
        else:
            return data_last_cpgt.strftime('%d.%m.%Y')

    @staticmethod
    def data_email():
        data_mes_email = dt.datetime.now() + relativedelta(months=1)
        data_email_1 = dt.datetime.now().strftime('01.%m.%Y')
        data_email_1_transforma_date = dt.datetime.strptime(data_email_1, '%d.%m.%Y')
        data_email_2 = dt.datetime.now().strftime('21.%m.%Y')
        data_email_2_transforma_date = dt.datetime.strptime(data_email_2, '%d.%m.%Y')
        data_email_3 = dt.datetime.now().strftime('%d.%m.%Y')
        data_email_3_transforma_date = dt.datetime.strptime(data_email_3, '%d.%m.%Y')
        data_email_1_day = data_email_1_transforma_date.day
        data_email_2_day = data_email_2_transforma_date.day
        data_email_3_day = data_email_3_transforma_date.day
        if data_email_3_day in range(0, (data_email_2_day - data_email_1_day)):
            return dt.datetime.now().strftime('%B')
        else:
            return data_mes_email.strftime('%B')

    def enviar_email(self):
        setlocale(LC_ALL, 'pt_BR.utf-8')
        pergunta_envio = input('Você deseja enviar o email agora?\n'
                               '(Pressione "enter" para enviar o email.\n'
                               'Caso deseje finalizar pressione "f" em seguida "enter".)-->')

        if pergunta_envio == '':
            data_visivel = (self.data_email()).title()

            smtpobj = smtplib.SMTP('smtp.gmail.com', 587)
            smtpobj.starttls()
            fro = 'jrf.@gmail.com'
            to = 'junio_@.com.br'

            smtpobj.login(fro, )
            msg = EmailMessage()
            msg['From'] = fro
            msg['To'] = to
            msg['Subject'] = f'Taxas Risco Sacado {data_visivel}.'

            msg.set_content(
                f'Prezada Elaine\n\nSegue abaixo as planilhas com as taxas dos clientes que utilizarão'
                f' as condições de pagamento na modalidade risco sacado para o mês de {data_visivel}.'
                f' Peço avaliar a solução.')
            paths = ['risco_sacado(' + self.data_save_arquivo() + ').xlsx',
                     'Cadastro_CPGT_RS(' + self.data_save_arquivo() + ').xlsx']
            for path in paths:
                caminho = open(path, 'rb')
                arq_data = caminho.read()
                arq_name = caminho.name
                msg.add_attachment(arq_data, maintype='application', subtype='octet-stream', filename=arq_name)

            smtpobj.send_message(msg)
            smtpobj.quit()
            print('Email enviado!!!!')
        elif pergunta_envio == "f":
            pass


x = Risco()
x.interface()


from openpyxl import load_workbook
import datetime as dt
from dateutil.relativedelta import relativedelta
import smtplib
from email.message import EmailMessage
from locale import setlocale, LC_ALL
import assists


class Risco:
    def __init__(self):
        self.cliente = 'utf-8'
        self.taxas = 0
        self.wb = load_workbook(filename='template_risco_sacado_new.xlsx')
        self.wb_cpgt = load_workbook(filename='template_cpgt_risco_sacado_new.xlsx')
        self.centro = int

    def interface(self):
        self.abrir_arq()
        active = True
        while active:
            self.cliente = input('Qual cliente você irá cadastrar? ').title()
            self.taxas = input('Qual a taxa? ')
            self.cpgt = input('Qual o prazo da condição de pagamento? ')
            self.banco = self.banco_1()
            self.lista_distr()
            self.abrir_plan_risco_sacado()
            self.abrir_plan_cpgt()
            self.salvar_arq()
            alerta = input('Prosseguir o cadastro ? \n(Pressione "enter" para continuar com os cadastros.\n'
                           'Caso deseje finalizar pressione "f" em seguida "enter".)-->')
            if alerta == 'f':
                active = False
        self.wb.close()
        self.wb_cpgt.close()
        #self.enviar_email()

    def abrir_arq(self):
        return self.wb and self.wb_cpgt

    def salvar_arq(self):
        self.wb.save('risco_sacado('+assists.data_cadastro()+').xlsx')
        self.wb_cpgt.save('Cadastro_em_lote_RS('+assists.data_cadastro()+').xlsx')

    def cpgt_terrestre(self):
        return 'ZD'+self.cpgt

    def cpgt_cabotagem(self):
        return 'ZC'+self.cpgt

    def cpgt_terrestre_cabotagem(self):
        if self.centro == 1401 or self.centro == 1211:
            return self.cpgt_cabotagem()
        else:
            return self.cpgt_terrestre()

    @staticmethod
    def banco_1():
        flag = True
        while flag:
            bancos = {'s': 'Santander', 'b': 'Bradesco', 'c': 'Citibank'}
            banco_marca = input('Escolha o banco?\n("s" para Santander, "b" para Bradesco e "c" para Citibank)'
                                ' + "enter" -->')
            if banco_marca in bancos:
                return bancos[banco_marca]
            else:
                print('Essa escolha não é possível, tente novamente!.')

    def lista_distr(self):
        distri = [self.cliente, self.taxas, self.cpgt_terrestre(), self.cpgt_cabotagem(), self.banco]
        return distri

    def clientes_distr(self):
        if self.lista_distr()[0] in self.distri_cliente_polo_produto():
            return self.lista_distr()[0]

    def distri_cliente_polo_produto(self):
        self.distribuidoras = {'Alesat': {8187: {1700: ['PB.620', 'PB.6DH', 'PB.658']},
                                      1740: {1400: ['PB.620', 'PB.6DH', 'PB.658']},
                                      4473: {1200: ['PB.620', 'PB.6DH', 'PB.658']},
                                      21699: {1100: ['PB.620', 'PB.6DH', 'PB.658']},
                                      4919: {1360: ['PB.6DH'], 1950: ['PB.6DH']},
                                      8429: {1101: ['PB.620', 'PB.6DH', 'PB.658']},
                                      #1733: {1110: ['PB.620', 'PB.6DH', 'PB.658']},
                                      #1732: {1111: ['PB.620', 'PB.6DH', 'PB.658']},
                                      #1736: {1120: ['PB.620', 'PB.6DH', 'PB.658']},
                                      6833: {1130: ['PB.620', 'PB.6DH', 'PB.658']}},
                               'Ciapetro': {455: {1400: ['PB.620', 'PB.6DH', 'PB.658']},
                                       18314: {1700: ['PB.620', 'PB.6DH', 'PB.658']},
                                       4150: {1100: ['PB.620', 'PB.6DH', 'PB.658']}},       #1110, 1111, 1120
                               'Ipp': {47: {1700: ['PB.620', 'PB.6DH', 'PB.658']},
                                       2093: {1400: ['PB.620', 'PB.6DH', 'PB.658']}},
                               'Mime': {17621: {1700: ['PB.620', 'PB.6DH', 'PB.658']}},
                               'Petrox': {5142: {1360: ['PB.6DH'], 1950: ['PB.6DH']}},
                               'Rodoil': {7008: {1700: ['PB.620', 'PB.6DH', 'PB.658']},
                                       6815: {1400: ['PB.620', 'PB.6DH', 'PB.658']}},
                               'Raizen': {49: {1700: ['PB.620', 'PB.6DH', 'PB.658']},
                                       2163: {1400: ['PB.620', 'PB.6DH', 'PB.658']},
                                       2153: {1200: ['PB.620', 'PB.6DH', 'PB.658'], 1210: ['PB.620', 'PB.6DH']},
                                       2150: {1100: ['PB.620', 'PB.6DH', 'PB.658']},
                                       2180: {1360: ['PB.6DH'], 1950: ['PB.6DH']},
                                       2155: {1101: ['PB.620', 'PB.6DH', 'PB.658']},
                                       #18449: {1110: ['PB.620', 'PB.6DH', 'PB.658']},
                                       #2186: {1111: ['PB.620', 'PB.6DH', 'PB.658']},
                                       #2168: {1120: ['PB.620', 'PB.6DH', 'PB.658']},
                                       2157: {1130: ['PB.620', 'PB.6DH', 'PB.658']}}}

        return self.distribuidoras

    @staticmethod
    def marca():
        return "x"

    @staticmethod
    def orgv():
        return "1001"

    @staticmethod
    def claros():
        return "02"

    @staticmethod
    def encargos():
        return '1,51% a.m'

    def tab(self):
        tabela = "665"
        return tabela

    def carencia_cpgt_terrestre(self):
        condicoes_cpgt = self.lista_distr()[2]
        list_separador = condicoes_cpgt.split('D')
        valor_separado = list_separador[1]
        resultado = int(valor_separado) - 1
        return resultado

    def carencia_cpgt_cabotagem(self):
        condicoes_cpgt_cabotagem = self.lista_distr()[3]
        list_separador_cabotagem = condicoes_cpgt_cabotagem.split('C')
        valor_separado_cabotagem = list_separador_cabotagem[1]
        resultado_cabotagem = int(valor_separado_cabotagem) - 4
        return resultado_cabotagem

    def carencia_cpgt_terrestre_cabotagem(self):
        if self.centro == 1401 or self.centro == 1211:
            return self.carencia_cpgt_cabotagem()
        else:
            return self.carencia_cpgt_terrestre()

    def abrir_plan_risco_sacado(self):
        aba_act = self.wb.active
        self.lista_distr()
        for linha_plan in range(aba_act.max_row + 1, aba_act.max_row + 2):
            info = self.distri_cliente_polo_produto()[self.lista_distr()[0]]
            for fili, info_1 in info.items():
                for self.centro, prod in info_1.items():
                    for combust in prod:
                        aba_act.cell(row=linha_plan, column=2).value = self.lista_distr()[0]
                        aba_act.cell(row=linha_plan, column=3).value = self.orgv()
                        aba_act.cell(row=linha_plan, column=4).value = self.claros()
                        aba_act.cell(row=linha_plan, column=5).value = fili
                        aba_act.cell(row=linha_plan, column=6).value = self.centro
                        aba_act.cell(row=linha_plan, column=7).value = combust
                        aba_act.cell(row=linha_plan, column=8).value = self.cpgt_terrestre_cabotagem()
                        aba_act.cell(row=linha_plan, column=9).value = self.lista_distr()[1]+' a.m.'
                        aba_act.cell(row=linha_plan, column=10).value = assists.data_inicio()
                        aba_act.cell(row=linha_plan, column=11).value = assists.data_last_day_risco_sacado()
                        aba_act.cell(row=linha_plan, column=12).value = self.lista_distr()[-1]
                        aba_act.cell(row=linha_plan, column=13).value = assists.data_cadastro()
                        aba_act.cell(row=linha_plan, column=14).value = self.encargos()
                        linha_plan += 1

    def abrir_plan_cpgt(self):
        aba_act_cpgt = self.wb_cpgt.active
        self.lista_distr()
        for linha_cpgt in range(aba_act_cpgt.max_row + 1, aba_act_cpgt.max_row + 2):
            info = self.distri_cliente_polo_produto()[self.lista_distr()[0]]
            for fili, info_1 in info.items():
                for self.centro, prod in info_1.items():
                    for combust in prod:
                        aba_act_cpgt.cell(row=linha_cpgt, column=1).value = self.marca()
                        aba_act_cpgt.cell(row=linha_cpgt, column=2).value = self.claros()
                        aba_act_cpgt.cell(row=linha_cpgt, column=4).value = self.orgv()
                        aba_act_cpgt.cell(row=linha_cpgt, column=7).value = self.carencia_cpgt_terrestre_cabotagem()
                        aba_act_cpgt.cell(row=linha_cpgt, column=8).value = self.centro
                        aba_act_cpgt.cell(row=linha_cpgt, column=9).value = combust
                        aba_act_cpgt.cell(row=linha_cpgt, column=10).value = fili
                        aba_act_cpgt.cell(row=linha_cpgt, column=12).value = "1"
                        aba_act_cpgt.cell(row=linha_cpgt, column=13).value = "BRL"
                        aba_act_cpgt.cell(row=linha_cpgt, column=14).value = "1"
                        aba_act_cpgt.cell(row=linha_cpgt, column=15).value = "M20"
                        aba_act_cpgt.cell(row=linha_cpgt, column=16).value = assists.data_inicio()
                        aba_act_cpgt.cell(row=linha_cpgt, column=17).value = assists.data_last_day_cpgt()
                        aba_act_cpgt.cell(row=linha_cpgt, column=18).value = self.tab()
                        aba_act_cpgt.cell(row=linha_cpgt, column=19).value = self.lista_distr()[0]
                        linha_cpgt += 1


x = Risco()
x.interface()
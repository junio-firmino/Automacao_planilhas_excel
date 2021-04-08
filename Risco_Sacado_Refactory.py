from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage
from locale import setlocale, LC_ALL
import assists


# I will refactory This project, for do it I chose the design pattern Facade.


class Managerriscosacado:
    print('-------------------------------------------------------------------\n'
          'Vamos iniciar o cadastro das condições do Risco Sacado para o Mês.')

    def __init__(self):
        self.plan_risco_sacado = None
        self.email = Email()
        self.contar_1 = 0

    def create_plan_risco_sacado(self):
        flag = True
        while flag:
            self.contar_1 += 1
            print(self.contar_1)
            listy = Listdistribuidora()
            if self.contar_1 > 1:
                listy.createlista_2()
            else:
                listy.createlista()

            listy0 = listy.list[0]  # template
            listy1 = listy.list[1]  # Cliente
            listy2 = listy.list[2]  # Taxas
            listy3 = listy.list[3]  # CPGT
            listy4 = listy.list[4]  # Banco
            self.plan_risco_sacado = Planriscosacado(template=listy0, cliente=listy1,
                                                     taxas=listy2, cpgt=listy3, banco=listy4)
            self.plan_risco_sacado.plan_taxes()
            self.plan_risco_sacado.plan_cpgt()
            alerta = input('-------------------------------------------------------------------\n'
                           'Prosseguir o cadastro ?\n'
                           '-------------------------------------------------------------------\n'
                           'Continuar --> "enter" \nFinalizar --> f + enter.-->')

            if alerta == 'f':  # tentar uma solução para a escolha dos templates para depois da pergunta acima por meio do IF
                flag = False
        self.email.enviar_email()


class Planriscosacado:
    def __init__(self, template, cliente, taxas, cpgt, banco):
        self.client0 = cliente
        self.taxas1 = taxas
        self.cpgt2 = cpgt
        self.banco3 = banco
        self.info = Informationconstant()
        self.carencia = Carencia(valores=Cpgt(cpgt))
        self.temp = template
        self.load = Loadworkbook(self.temp)
        self.cpgt_main = Cpgt(cpgt=cpgt)

    def plan_taxes(self):
        aba_act = self.load.wb.active
        for linha_plan in range(aba_act.max_row + 1, aba_act.max_row + 2):
            info = Distribuidoras().distri_cliente_polo_produto()[self.client0]
            for fili, info_1 in info.items():
                for centro, prod in info_1.items():
                    for combust in prod:
                        aba_act.cell(row=linha_plan, column=1).value = fili  # Filial
                        aba_act.cell(row=linha_plan, column=2).value = self.cpgt_main.cpgt_cabotagem() if \
                            centro == 1401 or centro == 1211 else self.cpgt_main.cpgt_terrestre()  # CPGT
                        aba_act.cell(row=linha_plan, column=3).value = combust  # Produto
                        aba_act.cell(row=linha_plan, column=4).value = centro  # Centro
                        aba_act.cell(row=linha_plan, column=6).value = self.taxas1 + ' a.m.'  # Taxas
                        aba_act.cell(row=linha_plan, column=7).value = "%"
                        aba_act.cell(row=linha_plan, column=10).value = "A"
                        aba_act.cell(row=linha_plan, column=12).value = assists.data_inicio()  # Data inicial
                        aba_act.cell(row=linha_plan, column=13).value = assists.data_last_day_risco_sacado()  # Data
                        aba_act.cell(row=linha_plan, column=14).value = self.client0  # Cliente
                        aba_act.cell(row=linha_plan, column=15).value = self.info.encargos()  # Encargos
                        aba_act.cell(row=linha_plan, column=16).value = self.banco3  # Banco
                        aba_act.cell(row=linha_plan, column=17).value = assists.data_cadastro()  # Data do cadastro
                        linha_plan += 1
        self.load.save()
        self.load.close()

    def plan_cpgt(self):
        aba_act_cpgt = self.load.open_wb_cpgt().active
        for linha_cpgt in range(aba_act_cpgt.max_row + 1, aba_act_cpgt.max_row + 2):
            info = Distribuidoras().distri_cliente_polo_produto()[self.client0]
            for fili, info_1 in info.items():
                for centro, prod in info_1.items():
                    for combust in prod:
                        aba_act_cpgt.cell(row=linha_cpgt, column=1).value = self.info.marca()
                        aba_act_cpgt.cell(row=linha_cpgt, column=2).value = self.info.claros()
                        aba_act_cpgt.cell(row=linha_cpgt, column=3).value = self.cpgt_main.cpgt_cabotagem() if \
                            centro == 1401 or centro == 1211 else self.cpgt_main.cpgt_terrestre()
                        aba_act_cpgt.cell(row=linha_cpgt, column=4).value = self.info.orgv()
                        aba_act_cpgt.cell(row=linha_cpgt,
                                          column=7).value = self.carencia.carencia_cpgt_cabotagem()if\
                            centro == 1401 or centro == 1211 else self.carencia.carencia_cpgt_terrestre()
                        aba_act_cpgt.cell(row=linha_cpgt, column=8).value = centro
                        aba_act_cpgt.cell(row=linha_cpgt, column=9).value = combust
                        aba_act_cpgt.cell(row=linha_cpgt, column=10).value = fili
                        aba_act_cpgt.cell(row=linha_cpgt, column=12).value = 1
                        aba_act_cpgt.cell(row=linha_cpgt, column=13).value = "BRL"
                        aba_act_cpgt.cell(row=linha_cpgt, column=14).value = 1
                        aba_act_cpgt.cell(row=linha_cpgt, column=15).value = "M20"
                        aba_act_cpgt.cell(row=linha_cpgt, column=16).value = "01.08.2020"
                        aba_act_cpgt.cell(row=linha_cpgt, column=17).value = "31.12.9999"
                        aba_act_cpgt.cell(row=linha_cpgt, column=18).value = self.info.tab()
                        aba_act_cpgt.cell(row=linha_cpgt, column=19).value = self.client0
                        linha_cpgt += 1
        self.load.save_wb_cpgt()
        self.load.close_wb_cpgt()


class Loadworkbook:
    def __init__(self, template):
        self.wb = load_workbook(filename=template + '.xlsx')
        self.wb_cpgt = load_workbook(filename='template_cpgt_' + template + '.xlsx')

    def open(self):
        return self.wb

    def open_wb_cpgt(self):
        return self.wb_cpgt

    def save(self):
        return self.wb.save('Risco Sacado - TMP(preço)(' + assists.data_cadastro() + ').xlsx')

    def save_wb_cpgt(self):
        return self.wb_cpgt.save('template_cpgt_Risco Sacado - TMP(preço)(' + assists.data_cadastro() + ').xlsx')

    def close(self):
        self.wb.close()

    def close_wb_cpgt(self):
        self.wb_cpgt.close()


class Listdistribuidora:
    def __init__(self):
        self.list = []

    def lista_distr(self, escolha):
        self.list.append(escolha)
        return self.list

    def createlista(self):   # A partir deste ponto o contador deve trabalhar
        self.lista_distr(Answer().template())
        self.lista_distr(Answer().client())
        self.lista_distr(Answer().taxas())
        self.lista_distr(Answer().cpgt())
        self.lista_distr(Answer().banco())

    def createlista_2(self):
        self.lista_distr(Answer().template_2())
        self.lista_distr(Answer().client())
        self.lista_distr(Answer().taxas())
        self.lista_distr(Answer().cpgt())
        self.lista_distr(Answer().banco())


class Cpgt:
    def __init__(self, cpgt):
        self.abs = cpgt

    def cpgt_terrestre(self):
        return 'ZD' + self.abs

    def cpgt_cabotagem(self):
        return 'ZC' + self.abs


class Informationconstant:
    @staticmethod
    def marca():
        return "x"

    @staticmethod
    def orgv():
        return 1001

    @staticmethod
    def claros():
        return "02"

    @staticmethod
    def encargos():
        return "1,51% a.m"

    @staticmethod
    def tab():
        tabela = 655
        return tabela


class Carencia:
    def __init__(self, valores):
        self.condicoes_cpgt = valores

    def carencia_cpgt_terrestre(self):
        list_separador = self.condicoes_cpgt.cpgt_terrestre().split('D')
        valor_separado = list_separador[1]
        resultado = int(valor_separado) - 1
        return resultado

    def carencia_cpgt_cabotagem(self):
        list_separador_cabotagem = self.condicoes_cpgt.cpgt_cabotagem().split('C')
        valor_separado_cabotagem = list_separador_cabotagem[1]
        resultado_cabotagem = int(valor_separado_cabotagem) - 4
        return resultado_cabotagem


class Distribuidoras:
    def __init__(self):
        self.distribuidoras = dict

    def distri_cliente_polo_produto(self):
        self.distribuidoras = {'Alesat': {8187: {1700: ['PB.620', 'PB.6DH', 'PB.658']},  # 1700 - Canoas
                                          1740: {1400: ['PB.620', 'PB.6DH', 'PB.658']},  # 1400 - Araucária
                                          4473: {1200: ['PB.620', 'PB.6DH', 'PB.658']},  # 1250 - Betim
                                          21699: {1100: ['PB.620', 'PB.6DH', 'PB.658']},  # 1200 - Cubatão
                                          4919: {1360: ['PB.6DH'], 1950: ['PB.6DH']},  # 1100 - Paulinia
                                          8429: {1101: ['PB.620', 'PB.6DH', 'PB.658']},  # 1101 - Ribeirão Preto
                                          # 1733: {1110: ['PB.620', 'PB.6DH', 'PB.658']},       # 1360 - Ipojuca
                                          # 1732: {1111: ['PB.620', 'PB.6DH', 'PB.658']},       # 1110 - Uberaba
                                          # 1736: {1120: ['PB.620', 'PB.6DH', 'PB.658']},  # 1111 - Uberlândia
                                          6833: {1130: ['PB.620', 'PB.6DH', 'PB.658']},  # 1120 - Senador Canedo
                                          6515: {1250: ['PB.620', 'PB.6DH', 'PB.658']}},  # 1130 - Brasília
                               'Ciapetro': {455: {1400: ['PB.620', 'PB.6DH', 'PB.658']},  # 1211 - Santos
                                            18314: {1700: ['PB.620', 'PB.6DH', 'PB.658']},  # 1401 - Paranaguá
                                            4150: {1100: ['PB.620', 'PB.6DH', 'PB.658']},  # 1110 - Uberaba*
                                            20497: {1250: ['PB.620', 'PB.6DH', 'PB.658']}},  # 1111 - Uberlandia*
                               'Ipp': {47: {1700: ['PB.620', 'PB.6DH', 'PB.658']},
                                       2093: {1400: ['PB.620', 'PB.6DH', 'PB.658']},
                                       2086: {1250: ['PB.620', 'PB.6DH', 'PB.658']},
                                       2102: {1250: ['PB.620', 'PB.6DH', 'PB.658']},
                                       15629: {1250: ['PB.620', 'PB.6DH', 'PB.658']}},
                               'Mime': {17621: {1700: ['PB.620', 'PB.6DH', 'PB.658']}},
                               'Petrox': {5142: {1360: ['PB.6DH'], 1950: ['PB.6DH']}},
                               'Rodoil': {7008: {1700: ['PB.6DH', 'PB.658']},
                                          6815: {1400: ['PB.6DH', 'PB.658']}},
                               'Raizen': {49: {1700: ['PB.620', 'PB.6DH', 'PB.658']},
                                          2163: {1400: ['PB.620', 'PB.6DH', 'PB.658']},
                                          2153: {1200: ['PB.620', 'PB.6DH', 'PB.658'], 1210: ['PB.620', 'PB.6DH']},
                                          2150: {1100: ['PB.620', 'PB.6DH', 'PB.658']},
                                          2180: {1360: ['PB.6DH'], 1950: ['PB.6DH']},
                                          2155: {1101: ['PB.620', 'PB.6DH', 'PB.658']},
                                          # 18449: {1110: ['PB.620', 'PB.6DH', 'PB.658']},
                                          # 2186: {1111: ['PB.620', 'PB.6DH', 'PB.658']},
                                          # 2168: {1120: ['PB.620', 'PB.6DH', 'PB.658']},
                                          2157: {1130: ['PB.620', 'PB.6DH', 'PB.658']},
                                          2144: {1250: ['PB.620', 'PB.6DH', 'PB.658']}},
                               'Rejaile': {19364: {1250: ['PB.6DH', 'PB.658']},
                                           21184: {1700: ['PB.6DH', 'PB.658']},
                                           156: {1400: ['PB.6DH', 'PB.658']}},
                               'Total': {21973: {1250: ['PB.620', 'PB.6DH', 'PB.658']},
                                         22176: {1130: ['PB.620', 'PB.6DH', 'PB.658']},
                                         21997: {1101: ['PB.620', 'PB.6DH', 'PB.658']}},
                               'Rio Branco': {8235: {1100: ['PB.620', 'PB.6DH', 'PB.658']},
                                              15456: {1101: ['PB.620', 'PB.6DH', 'PB.658']},
                                              5218: {1250: ['PB.620', 'PB.6DH', 'PB.658']}}}
        return self.distribuidoras


class Email:
    @staticmethod
    def enviar_email():
        setlocale(LC_ALL, 'pt_BR.utf-8')
        pergunta_envio = input('-------------------------------------------------------------------\n'
                               'Você deseja enviar o email agora?\n'
                               '-------------------------------------------------------------------\n'
                               '"enter" para enviar ou "f" + "enter" para finalizar.--> ')

        if pergunta_envio == '':
            data_visivel = assists.data_email().title()

            smtpobj = smtplib.SMTP('smtp.gmail.com', 587)
            smtpobj.starttls()
            fro = ''
            to = ''

            smtpobj.login(fro, '')
            msg = EmailMessage()
            msg['From'] = fro
            msg['To'] = to
            msg['Subject'] = f'Taxas Risco Sacado {data_visivel}.'

            msg.set_content(
                f'Prezados\n\nSegue abaixo a planilha com as taxas dos clientes que utilizarão'
                f' as condições de pagamento na modalidade risco sacado para o mês de {data_visivel}.')
            # trabalhar neste ponto da escolha dos arquivos
            paths = ['Risco Sacado - TMP(preço)(' + assists.data_cadastro() + ').xlsx',
                     'template_cpgt_Risco Sacado - TMP(preço)(' + assists.data_cadastro() + ').xlsx']
            for path in paths:
                caminho = open(path, 'rb')
                arq_data = caminho.read()
                arq_name = caminho.name
                msg.add_attachment(arq_data, maintype='application', subtype='octet-stream', filename=arq_name)

            smtpobj.send_message(msg)
            smtpobj.quit()
            print('Email enviado!!!!')
        elif pergunta_envio == "f":
            pass


class Answer:
    def __init__(self):
        self.ask_cliente_distr = str
        self.taxas_1 = str
        self.banco_choice = str
        self.cpgt_1 = str

    @staticmethod
    def template():
        flag_template = True
        while flag_template:
            arquivo = input('-------------------------------------------------------------------\n'
                            '1 - NOVO ARQUIVO \n'
                            '2 - Arquivo atual\n'
                            '3 - Escolha somente o dia deste mês do arquivo\n'  
                            '-------------------------------------------------------------------\n'
                            'Escolha das opções acima qual tipo de arquivo utilizar ---> ')

            arquivos = {'1': 'Risco Sacado - TMP(preço)', '2': 'Risco Sacado - TMP(preço)'
                                                               '(' + assists.data_cadastro() + ')'}
            if arquivo == '' or arquivo != '1' and arquivo != '2' and arquivo != '3':
                print('-------------------------------------------------------------------\n'
                      '-------------------------------------------------------------------\n'
                      'Não é possível utilizar essa escolha, tente alguma das'
                      ' opções de\ntemplates  abaixo para prosseguir.\n'
                      '-------------------------------------------------------------------')

            elif arquivo == '3':
                answer_1 = input('Qual o dia do mês corrente do arquivo você quer trabalhar? --> ')
                return 'Risco Sacado - TMP(preço)(' + answer_1 + assists.data_cadastro_month() + ')'

            elif arquivo in arquivos:
                choice_arquivos = arquivos[arquivo]
                return choice_arquivos

            else:
                return arquivo

    @staticmethod
    def template_2():
        return 'Risco Sacado - TMP(preço)(' + assists.data_cadastro() + ')'

    def client(self):
        flag_cli = True
        while flag_cli:
            cliente_distr = ['Alesat', 'Ciapetro', 'Ipp', 'Mime', 'Petrox', 'Rodoil', 'Raizen', 'Rejaile',
                             'Total', 'Rio Branco']
            self.ask_cliente_distr = input('Qual cliente você irá cadastrar? ').title()
            if self.ask_cliente_distr in cliente_distr:
                return self.ask_cliente_distr
            else:
                print('Empresa não participante do Risco Sacado, tente outra empresa!.')

    def taxas(self):
        self.taxas_1 = input('Qual a taxa? ')
        return self.taxas_1

    def cpgt(self):
        self.cpgt_1 = input('Qual é a condição de pagamento? ')
        return self.cpgt_1

    def banco(self):
        flag = True
        while flag:
            bancos = {'s': 'Santander', 'b': 'Bradesco', 'c': 'Citibank'}
            banco_marca = input('Escolha o banco?\n("s" para Santander, "b" para Bradesco e "c" para Citibank)'
                                ' + "enter"-->')
            if banco_marca in bancos:
                self.banco_choice = bancos[banco_marca]
                return self.banco_choice
            else:
                print('-------------------------------------------------------------------\n'
                      'Essa escolha não é possível, tente novamente!.\n'
                      '-------------------------------------------------------------------')


class Interface:
    @staticmethod
    def askinterface():
        me = Managerriscosacado()
        return me.create_plan_risco_sacado()


if __name__ == '__main__':
    inte = Interface()
    inte.askinterface()

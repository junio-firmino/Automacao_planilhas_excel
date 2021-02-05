from openpyxl import load_workbook
import datetime as dt
from dateutil.relativedelta import relativedelta
import assists


class Parametros:
    def __init__(self):
        self.montante_cgv = 0
        self.montante_a = 0
        self.montante_sp = 0
        self.montante_a_n4 = 0
        self.montante_sp_n4 = 0
        self.montante_f1 = 0
        self.montante_f2 = 0
        self.montante_f3 = 0
        self.ask = str
        self.client = dict
        self.condicoes_parametro = list
        self.wb = load_workbook(filename='template_PVA_PVS.xlsx')
        self.wb_n4 = load_workbook(filename='template_PVA_PVS.xlsx')
        self.wb_cgv = load_workbook(filename='template_PVA_PVS.xlsx')
        self.wb_zav = load_workbook(filename='template_ZMUE.xlsx')
        self.wb_zn4 = load_workbook(filename='template_ZMUE.xlsx')
        self.pergunta_1 = None
        self.produto_centro = None
        self.cliente_matriz = None
        self.open_arq = None
        self.condicoes_parametro_cgv = None
        self.condicoes_parametro_zavulso = None

    def interface_client(self):
        flag = True
        while flag:
            self.pergunta_1 = self.pergunta()
            self.abrir_arq()
            self.montante()
            self.planilha_referente_contrato()
            self.save_arq()
            self.fechar_arq()
            alert = input('Prosseguir o cadastro ? \n(Pressione "enter" para continuar com os cadastros.\n'
                          'Caso deseje finalizar selecione "f" em seguida "enter".)-->')
            if alert == 'f':
                flag = False
        print('Processo finalizado!.')

    def pergunta(self):
        contrato_escolhido = ['Cgv', 'N4', 'Avulso', 'Zavulso', 'Zn4']
        active = True
        while active:
            self.ask = input("Escolha o tipo de contrato -->  ").title()
            if self.ask in contrato_escolhido:
                return self.ask
            else:
                print("Essa escolha não é possível, tente novamente!.")

    def planilha_referente_contrato(self):
        if self.list_trabalho()[0] == 'Avulso':
            return self.planilha_avulso()
        if self.list_trabalho()[0] == 'Cgv':
            return self.planilha_cgv()
        if self.list_trabalho()[0] == 'N4':
            return self.planilha_n4()
        if self.list_trabalho()[0] == 'Zavulso':
            return self.planilha_zavulso()
        if self.list_trabalho()[0] == 'Zn4':
            return self.planilha_zn4()

    def montante(self):
        if self.list_trabalho()[0] == 'Avulso' or self.list_trabalho()[0] == 'N4':
            self.montante_a = input('Qual o parâmetro para o Adicional (PVA)?')
            self.montante_sp = input('Qual o parâmetro para o Suplementar (PVS)?')
        else:
            self.montante_a = 0
            self.montante_sp = 0

        if self.list_trabalho()[0] == 'Cgv':
            self.montante_cgv = input('Qual o parâmetro CGV para o Adicional (PVA)? ')
        else:
            self.montante_cgv = 0

        if self.list_trabalho()[0] == 'Zavulso' or self.list_trabalho()[0] == 'Zn4':
            self.montante_f1 = input('Qual o parâmetro ZMUE para 0% < faixa < 5%? ')
            self.montante_f2 = input('Qual o parâmetro ZMUE para 5% < faixa < 10%? ')
            self.montante_f3 = input('Qual o parâmetro ZMUE para a faixa > 10%? ')
        else:
            self.montante_f1 = 0
            self.montante_f2 = 0
            self.montante_f3 = 0

    def list_trabalho(self):
        work = [self.pergunta_1, self.montante_cgv, self.montante_a, self.montante_sp, self.montante_f1,
                self.montante_f2, self.montante_f3]
        return work

    def cliente_centro_produto(self):
        self.client = {
            # 15640: {'PB.620': [1100, 1101, 1150, 1160]},  # Rumos
            # 725: {'PB.620': [1100, 1101], 'PB.658': [1100, 1101, 1160], 'PB.6DH': [1160]},
            # 20347: {'PB.620': [1100], 'PB.658': [1100]},
            17644: {'PB.620': [1101], 'PB.658': [1101]},  # Torrão
            # 4168: {'PB.620': [1111], 'PB.658': [1111]},
            21254: {'PB.620': [1120]},  # Rumos
            16906: {'PB.620': [1120], 'PB.658': [1120], 'PB.6DH': [1120]},  # Tabocão
            724: {'PB.620': [1120], 'PB.658': [1120], 'PB.6DH': [1120]},  # Tabocão
            # 8425: {'PB.620': [1150]},
            # 15630: {'PB.620': [1150], 'PB.658': [1150], 'PB.6DH': [1150]},
            # 766: {'PB.620': [1160]}, # Aster (760)
            766: {'PB.658': [1160], 'PB.6DH': [1160]},  # Aster (760)
            # 21933: {'PB.620': [1160, 1150]},
            # 8944: {'PB.620': [1160], 'PB.658': [1160], 'PB.6DH': [1160]},
            # 17984: {'PB.620': [1250], 'PB.658': [1250], 'PB.6DH': [1250]},
            # 16350: {'PB.620': [1360, 9044], 'PB.658': [1360, 9044], 'PB.6DH': [1360]},
            1123: {'PB.658': [1500, 1507], 'PB.6DH': [1500, 1507], 'PB.650': [1500]},  # Sabba
            1125: {'PB.658': [1502, 9102], 'PB.6DH': [1502, 9102], 'PB.650': [1502]},  # Sabba
            1124: {'PB.658': [1560, 9842, 9846, 9848], 'PB.6DH': [1560, 9842, 9848]},  # Sabba
            10174: {'PB.658': [1560, 9848], 'PB.6DH': [1560, 9848]},  # 'PB.620': [1560, 9848]},  # Tabocão
            7169: {'PB.658': [1506]},  # Sabba
            # 6697: {'PB.658': [9842], 'PB.6DH': [9842]},
            4432: {'PB.650': [1050]},  # Tobras
            156: {'PB.620': [1400], 'PB.650': [1400]},  # Rejaile
            # 157: {'PB.650': [1400]},
            155: {'PB.620': [1423], 'PB.650': [1423]},  # Rejaile
            18334: {'PB.620': [1050]},  # Rejaile
            18439: {'PB.620': [1111]},  # Rejaile
            19364: {'PB.620': [1250]},  # Rejaile
            21184: {'PB.620': [1700]},  # Rejaile
            10334: {'PB.620': [1700]},  # Rejaile
            10132: {'PB.620': [1421]},  # Rodoil
            21665: {'PB.620': [1422]},  # Rodoil
            6815: {'PB.620': [1400]},  # Rodoil
            7008: {'PB.620': [1700]},  # Rodoil
        }
        return self.client

    def produto_centro_cgv(self):
        self.produto_centro = {'PB.620': [1160, 1423, 1362, 1350, 1422, 1400, 1110, 1352, 1550, 1150, 1111, 1502,
                                          1365, 1560, 2540, 1100, 1050, 1360, 1101, 1421, 1700, 1250, 1120, 1070,
                                          1312, 1130, 9060, 1353, 1200, 1500, 1354, 1507, 1311, 1062, 1710],
                               'PB.650': [1500, 1400, 1423, 1050, 1560, 1200, 1550, 1502, 2540, 1365, 9060, 1070,
                                          1350, 1362, 1422],
                               'PB.658': [1160, 1423, 1362, 1350, 1422, 1400, 1110, 1352, 1550, 1150, 1111, 1502,
                                          1365, 1560, 2540, 1100, 1050, 1360, 1101, 1421, 1700, 1250, 1120, 1070,
                                          1312, 1130, 9060, 1353, 1200, 1500, 1354, 1507, 1311, 1062, 1710],
                               'PB.6DH': [1160, 1423, 1362, 1350, 1422, 1400, 1110, 1352, 1550, 1150, 1111, 1502,
                                          1365, 1560, 2540, 1100, 1050, 1360, 1101, 1421, 1700, 1250, 1120, 1070,
                                          1312, 1130, 9060, 1353, 1200, 1500, 1354, 1507, 1311, 1062, 1710]}
        return self.produto_centro

    def cliente_matriz_z(self):
        self.cliente_matriz = {'PB.620': [760, 623, 724, 154, 6814, 15640, 6997],
                               'PB.650': [1122, 154, 4432],
                               'PB.658': [760, 623, 724, 1122, 6997],
                               'PB.6DH': [760, 623, 724, 1122, 6997]}
        return self.cliente_matriz

    def abrir_arq(self):
        self.open_arq = {'Avulso': self.wb, 'N4': self.wb_n4, 'Cgv': self.wb_cgv,
                         'Zavulso': self.wb_zav, 'Zn4': self.wb_zn4}
        return self.open_arq[self.list_trabalho()[0]]

    def save_arq(self):
        salvar = self.open_arq
        if self.list_trabalho()[0] == 'Avulso' or self.list_trabalho()[0] == 'N4':
            return salvar[self.list_trabalho()[0]].save(
                'PVA_PVS_' + (self.list_trabalho()[0]).upper() + '(' + assists.data_cadastro() + ')' + '.xlsx')
        elif self.list_trabalho()[0] == 'Cgv':
            return salvar[self.list_trabalho()[0]].save(
                'PVA_' + (self.list_trabalho()[0]).upper() + '(' + assists.data_cadastro() + ')' + '.xlsx')
        elif self.list_trabalho()[0] == 'Zavulso' or self.list_trabalho()[0] == 'Zn4':
            return salvar[self.list_trabalho()[0]].save(
                (self.list_trabalho()[0]).upper() + '_' + '(' + assists.data_cadastro() + ').xlsx')

    def fechar_arq(self):
        close = self.open_arq
        return close[self.list_trabalho()[0]].close()

    @staticmethod
    def marca():
        return "x"

    @staticmethod
    def claros():
        return "02"

    @staticmethod
    def orgv():
        return "1001"

    def tipo_contrato(self):
        contract = {'Avulso': 'N4', 'N4': 'N4', 'Cgv': 'P'}
        return contract[self.list_trabalho()[0]]

    def grc4(self):
        if self.list_trabalho()[0] == 'Avulso' or self.list_trabalho()[0] == 'N4':
            self.condicoes_parametro = {'A': self.list_trabalho()[2], 'SP': self.list_trabalho()[3]}
            return self.condicoes_parametro

        if self.list_trabalho()[0] == 'Cgv':
            self.condicoes_parametro_cgv = {'A': self.list_trabalho()[1]}
            return self.condicoes_parametro_cgv

        if self.list_trabalho()[0] == 'Zavulso' or self.list_trabalho()[0] == 'Zn4':
            self.condicoes_parametro_zavulso = {5: self.list_trabalho()[4], 10: self.list_trabalho()[5],
                                                100: self.list_trabalho()[6]}
            return self.condicoes_parametro_zavulso

    @staticmethod
    def material():
        produto = ['PB.620', 'PB.6DH', 'PB.658', 'PB.650']
        return produto

    @staticmethod
    def moeda():
        return 'BRL'

    @staticmethod
    def por():
        return "1"

    @staticmethod
    def unidade():
        return "M20"

    def tab(self):
        tabela = {'Cgv': 689, 'Avulso': 525, 'N4': 556, 'Zavulso': 652, 'Zn4': 669}
        return tabela[self.list_trabalho()[0]]

    @staticmethod
    def data_inicial():
        data_inicio = dt.datetime.now() + relativedelta(day=1, months=1)
        data_inicio_return = data_inicio.strftime('%d.%m.%Y')
        return data_inicio_return

    @staticmethod
    def data_inicial_z():
        data_inicio = dt.datetime.now() + relativedelta(day=1, months=2)
        data_inicio_return = data_inicio.strftime('%d.%m.%Y')
        return data_inicio_return

    @staticmethod
    def data_fim():
        data_last = dt.datetime.now() + relativedelta(day=31, months=1)
        data_return = data_last.strftime('%d.%m.%Y')
        return data_return

    @staticmethod
    def data_fim_z():
        data_last = dt.datetime.now() + relativedelta(day=31, months=2)
        data_return = data_last.strftime('%d.%m.%Y')
        return data_return

    def planilha_avulso(self):
        aba_avulso = self.wb.active
        self.list_trabalho()
        for linha_plan in range(3, 4):
            for condi, valor in self.grc4().items():
                for filiais, carac in self.cliente_centro_produto().items():
                    for product, centre in carac.items():
                        for numero_centre in centre:
                            aba_avulso.cell(row=linha_plan, column=1).value = self.marca()
                            aba_avulso.cell(row=linha_plan, column=2).value = self.claros()
                            aba_avulso.cell(row=linha_plan, column=3).value = self.orgv()
                            aba_avulso.cell(row=linha_plan, column=6).value = self.tipo_contrato()
                            aba_avulso.cell(row=linha_plan, column=12).value = valor
                            aba_avulso.cell(row=linha_plan, column=13).value = self.moeda()
                            aba_avulso.cell(row=linha_plan, column=14).value = self.por()
                            aba_avulso.cell(row=linha_plan, column=15).value = self.unidade()
                            aba_avulso.cell(row=linha_plan, column=16).value = self.data_inicial()
                            aba_avulso.cell(row=linha_plan, column=17).value = self.data_fim()
                            aba_avulso.cell(row=linha_plan, column=18).value = self.tab()
                            aba_avulso.cell(row=linha_plan, column=4).value = condi
                            aba_avulso.cell(row=linha_plan, column=8).value = filiais
                            aba_avulso.cell(row=linha_plan, column=9).value = product
                            aba_avulso.cell(row=linha_plan, column=7).value = numero_centre
                            linha_plan += 1

    def planilha_n4(self):
        aba_n4 = self.wb_n4.active
        self.list_trabalho()
        for linha_plan in range(3, 4):
            for condi_n4, valor_n4 in self.grc4().items():
                for gas in self.material():
                    aba_n4.cell(row=linha_plan, column=1).value = self.marca()
                    aba_n4.cell(row=linha_plan, column=2).value = self.claros()
                    aba_n4.cell(row=linha_plan, column=3).value = self.orgv()
                    aba_n4.cell(row=linha_plan, column=4).value = condi_n4
                    aba_n4.cell(row=linha_plan, column=9).value = gas
                    aba_n4.cell(row=linha_plan, column=12).value = valor_n4
                    aba_n4.cell(row=linha_plan, column=13).value = self.moeda()
                    aba_n4.cell(row=linha_plan, column=14).value = self.por()
                    aba_n4.cell(row=linha_plan, column=15).value = self.unidade()
                    aba_n4.cell(row=linha_plan, column=16).value = self.data_inicial()
                    aba_n4.cell(row=linha_plan, column=17).value = self.data_fim()
                    aba_n4.cell(row=linha_plan, column=18).value = self.tab()
                    linha_plan += 1

    def planilha_cgv(self):
        aba_cgv = self.wb_cgv.active
        self.list_trabalho()
        for linha_plan in range(3, 4):
            for condicao, valorr in self.grc4().items():
                for producto, centro in self.produto_centro_cgv().items():
                    for numero_centro in centro:
                        aba_cgv.cell(row=linha_plan, column=1).value = self.marca()
                        aba_cgv.cell(row=linha_plan, column=2).value = self.claros()
                        aba_cgv.cell(row=linha_plan, column=3).value = self.orgv()
                        aba_cgv.cell(row=linha_plan, column=6).value = self.tipo_contrato()
                        aba_cgv.cell(row=linha_plan, column=12).value = valorr
                        aba_cgv.cell(row=linha_plan, column=13).value = self.moeda()
                        aba_cgv.cell(row=linha_plan, column=14).value = self.por()
                        aba_cgv.cell(row=linha_plan, column=15).value = self.unidade()
                        aba_cgv.cell(row=linha_plan, column=16).value = self.data_inicial()
                        aba_cgv.cell(row=linha_plan, column=17).value = self.data_fim()
                        aba_cgv.cell(row=linha_plan, column=18).value = self.tab()
                        aba_cgv.cell(row=linha_plan, column=4).value = condicao
                        aba_cgv.cell(row=linha_plan, column=9).value = producto
                        aba_cgv.cell(row=linha_plan, column=7).value = numero_centro
                        linha_plan += 1

    def planilha_zavulso(self):
        aba_zavulso = self.wb_zav.active
        self.list_trabalho()
        for linha_plan in range(3, 4):
            for product, matriz in self.cliente_matriz_z().items():
                for matrizes in matriz:
                    for escala, valor in self.grc4().items():
                        aba_zavulso.cell(row=linha_plan, column=1).value = self.marca()
                        aba_zavulso.cell(row=linha_plan, column=2).value = self.claros()
                        aba_zavulso.cell(row=linha_plan, column=3).value = self.orgv()
                        aba_zavulso.cell(row=linha_plan, column=4).value = 'N4'
                        aba_zavulso.cell(row=linha_plan, column=9).value = matrizes
                        aba_zavulso.cell(row=linha_plan, column=10).value = product
                        aba_zavulso.cell(row=linha_plan, column=11).value = valor
                        aba_zavulso.cell(row=linha_plan, column=12).value = self.moeda()
                        aba_zavulso.cell(row=linha_plan, column=13).value = self.por()
                        aba_zavulso.cell(row=linha_plan, column=14).value = self.unidade()
                        aba_zavulso.cell(row=linha_plan, column=15).value = self.data_inicial_z()
                        aba_zavulso.cell(row=linha_plan, column=16).value = self.data_fim_z()
                        aba_zavulso.cell(row=linha_plan, column=17).value = escala
                        aba_zavulso.cell(row=linha_plan, column=18).value = self.moeda()
                        aba_zavulso.cell(row=linha_plan, column=19).value = self.tab()
                        linha_plan += 1

    def planilha_zn4(self):
        aba_zn4 = self.wb_zn4.active
        self.list_trabalho()
        for linha_plan in range(3, 4):
            for gas in self.material():
                for condi_n4, valor_n4 in self.grc4().items():
                    aba_zn4.cell(row=linha_plan, column=1).value = self.marca()
                    aba_zn4.cell(row=linha_plan, column=2).value = self.claros()
                    aba_zn4.cell(row=linha_plan, column=3).value = self.orgv()
                    aba_zn4.cell(row=linha_plan, column=4).value = 'N4'
                    aba_zn4.cell(row=linha_plan, column=10).value = gas
                    aba_zn4.cell(row=linha_plan, column=11).value = valor_n4
                    aba_zn4.cell(row=linha_plan, column=12).value = self.moeda()
                    aba_zn4.cell(row=linha_plan, column=13).value = self.por()
                    aba_zn4.cell(row=linha_plan, column=14).value = self.unidade()
                    aba_zn4.cell(row=linha_plan, column=15).value = self.data_inicial_z()
                    aba_zn4.cell(row=linha_plan, column=16).value = self.data_fim_z()
                    aba_zn4.cell(row=linha_plan, column=17).value = condi_n4
                    aba_zn4.cell(row=linha_plan, column=18).value = self.moeda()
                    aba_zn4.cell(row=linha_plan, column=19).value = self.tab()
                    linha_plan += 1


x = Parametros()
x.interface_client()
